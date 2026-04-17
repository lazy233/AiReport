"""从任意表头的 CSV/Excel 批量导入学生数据：大模型将行映射到系统 profile 白名单字段。"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
from typing import Any

from ppt_report.models import db as db_mod
from ppt_report.services.dashscope_client import dashscope_chat_completion
from ppt_report.services.llm_json import extract_json_from_text

log = logging.getLogger(__name__)

# 每批行数：控制单次模型输出大小与稳定性；10 条在吞吐与可靠性之间较均衡
STUDENT_IMPORT_BATCH_SIZE = max(1, int(os.getenv("STUDENT_IMPORT_BATCH_SIZE", "10")))

_CELL_MAX = 800

# 与前端 student-data-store.js PROFILE_SPEC 对齐，仅允许这些键写入 profile
_PROFILE_DIM_KEYS: dict[str, frozenset[str]] = {
    "basic": frozenset(
        {
            "studentName",
            "nicknameEn",
            "school",
            "major",
            "gradeLevel",
            "gradeIntake",
            "currentTerm",
            "product",
            "serviceStart",
            "plannerTeacher",
            "advisorTeacher",
            "studentId",
            "className",
            "email",
            "phone",
            "remark",
        },
    ),
    "learning": frozenset(
        {
            "strength_subjects",
            "scores",
            "learning_good",
            "learning_weak",
            "interests",
            "study_goal",
            "career_goal",
            "long_goal",
            "degree",
            "duration",
            "credits",
            "course_rule",
            "gpa_rule",
            "selection_rule",
            "recommended_courses",
            "course_notes",
            "term_plan",
            "future_plan",
        },
    ),
    "hours": frozenset(
        {
            "totalHours",
            "usedHours",
            "remainingHours",
            "prep_courses",
            "tutoring_courses",
            "skillDirection",
            "skillDescription",
        },
    ),
    "guidance": frozenset(
        {
            "termSummary",
            "courseFeedback",
            "shortTermAdvice",
            "longTermDevelopment",
        },
    ),
    "term_summary": frozenset(
        {
            "student_summary",
            "school_ddl",
            "first_class_time",
            "first_class_note",
            "summer_work",
            "term_work",
            "recorded_courses",
            "grades",
            "gpa",
            "target_gpa",
            "final_score",
            "services",
            "service_count",
            "class_count",
            "total_duration",
            "avg_duration",
            "communication",
            "next_goal",
            "risk_courses",
            "suggestions",
            "remarks",
        },
    ),
}


def _schema_json_for_prompt() -> str:
    lines: list[str] = []
    dim_titles = {
        "basic": "基础信息",
        "learning": "学习画像",
        "hours": "课时数据",
        "guidance": "成长指导数据",
        "term_summary": "学期总结/结单",
    }
    for dim, keys in _PROFILE_DIM_KEYS.items():
        lines.append(f"【{dim_titles.get(dim, dim)}】维度 id=\"{dim}\"，字段（key: 含义略）: {sorted(keys)}")
    return "\n".join(lines)


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if len(s) > _CELL_MAX:
        s = s[: _CELL_MAX - 1] + "…"
    return s


def _parse_csv(raw: bytes) -> tuple[list[str], list[list[str]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows_raw = [list(r) for r in reader]
    if not rows_raw:
        return [], []
    width = max(len(r) for r in rows_raw)
    rows = [r + [""] * (width - len(r)) for r in rows_raw]
    headers = [_norm_cell(h) for h in rows[0]]
    data = [[_norm_cell(c) for c in row] for row in rows[1:]]
    data = [row + [""] * (width - len(row)) for row in data]
    return headers, data


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("未安装 openpyxl，无法读取 .xlsx，请安装依赖：pip install openpyxl") from exc

    bio = io.BytesIO(raw)
    wb = load_workbook(filename=bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_rows: list[list[Any]] = []
        for row in rows_iter:
            raw_rows.append(list(row))
    finally:
        wb.close()
    if not raw_rows:
        return [], []
    width = max(len(r) for r in raw_rows)
    padded = [list(r) + [None] * (width - len(r)) for r in raw_rows]
    headers = [_norm_cell(padded[0][j]) for j in range(width)]
    data = []
    for r in padded[1:]:
        data.append([_norm_cell(r[j]) for j in range(width)])
    return headers, data


def parse_spreadsheet(raw: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return _parse_csv(raw)
    if fn.endswith(".xlsx"):
        return _parse_xlsx(raw)
    raise ValueError("仅支持 .csv 或 .xlsx 文件。")


def _sanitize_profile_section(dim: str, src: Any) -> dict[str, str]:
    allowed = _PROFILE_DIM_KEYS.get(dim, frozenset())
    if not isinstance(src, dict):
        return {}
    out: dict[str, str] = {}
    for k in allowed:
        if k not in src:
            continue
        out[k] = _norm_cell(src.get(k))
    return out


def _sanitize_profile(raw: Any) -> dict[str, dict[str, str]]:
    if not isinstance(raw, dict):
        raw = {}
    return {
        dim: _sanitize_profile_section(dim, raw.get(dim))
        for dim in _PROFILE_DIM_KEYS.keys()
    }


def _record_to_save_payload(rec: dict[str, Any]) -> dict[str, Any]:
    profile = _sanitize_profile(rec.get("profile"))
    name = _norm_cell(rec.get("name"))
    sid = _norm_cell(rec.get("studentId"))
    basic = dict(profile.get("basic") or {})
    if not name:
        name = _norm_cell(basic.get("studentName"))
    if not sid:
        sid = _norm_cell(basic.get("studentId"))
    if name:
        basic["studentName"] = name
    if sid:
        basic["studentId"] = sid
    profile["basic"] = basic
    return {
        "name": name,
        "studentId": sid,
        "profile": profile,
        "content": _norm_cell(rec.get("content")),
    }


def _call_llm_map_rows(
    headers: list[str],
    rows: list[list[str]],
    batch_index: int,
    batch_total: int,
) -> list[dict[str, Any]]:
    payload_table = {"headers": headers, "rows": rows}
    system = (
        "你是教育数据结构化助手。用户上传的表格列名和顺序不固定，可能与系统字段不完全一致。"
        "请根据列名与单元格内容的语义，将每一行映射为一条学生档案。\n"
        "规则：\n"
        "1) 只输出一个 JSON 对象，不要 Markdown 代码围栏，不要其它文字。\n"
        '2) 顶层必须为 {"records": [ ... ] }，records 数组长度必须等于输入行数（本批 '
        f"{len(rows)} 行）。\n"
        "3) records[i] 对应输入的第 i 行数据（0 起始）。\n"
        "4) 每条记录结构："
        '{"name":"姓名或空", "studentId":"学号或空", "content":"数据内容正文或空", '
        '"profile": { "basic":{}, "learning":{}, "hours":{}, "guidance":{}, "term_summary":{} } }\n'
        "5) profile 各维度只能使用下方 SCHEMA 中出现的英文键名；无法从表格推断的字段置为 \"\" 或省略。\n"
        "6) 禁止编造：不得填充输入表格中完全不存在的具体事实（分数、学校名等）；不知道的留空。\n"
        "7) 若某行明显不是学生数据且无法解析，仍输出占位记录，尽量从行中提取姓名/学号，其余为空。"
    )
    user = (
        "系统允许的字段 SCHEMA（键名必须一致）：\n"
        f"{_schema_json_for_prompt()}\n\n"
        f"当前为第 {batch_index + 1}/{batch_total} 批，本批共 {len(rows)} 行。\n\n"
        "表格 JSON：\n"
        f"{json.dumps(payload_table, ensure_ascii=False)}"
    )
    raw_content = dashscope_chat_completion(
        system=system,
        user=user,
        temperature=0.1,
        max_tokens=None,
        timeout_sec=None,
    )
    try:
        parsed = extract_json_from_text(raw_content)
    except json.JSONDecodeError as exc:
        raise RuntimeError("模型未返回合法 JSON，请重试或减小文件行数后再试。") from exc
    if not isinstance(parsed, dict):
        raise RuntimeError("模型返回不是 JSON 对象。")
    records = parsed.get("records")
    if not isinstance(records, list):
        raise RuntimeError("模型返回缺少 records 数组。")
    if len(records) != len(rows):
        log.warning(
            "import_ai: batch records count mismatch: got %s want %s",
            len(records),
            len(rows),
        )
    out: list[dict[str, Any]] = []
    for i, item in enumerate(records):
        if isinstance(item, dict):
            out.append(item)
        else:
            out.append({})
    while len(out) < len(rows):
        out.append({})
    return out[: len(rows)]


def _row_nonempty(row: list[str]) -> bool:
    return any(bool(_norm_cell(c)) for c in row)


def run_smart_import(file_bytes: bytes, filename: str) -> dict[str, Any]:
    if not db_mod.db_enabled():
        raise RuntimeError("数据库未启用。")
    headers, data_rows = parse_spreadsheet(file_bytes, filename)
    if not headers:
        raise ValueError("文件中没有表头或内容为空。")
    data_rows = [r for r in data_rows if _row_nonempty(r)]
    if not data_rows:
        raise ValueError("没有可导入的数据行（除表头外全为空）。")

    n = len(data_rows)
    batch_size = STUDENT_IMPORT_BATCH_SIZE
    num_batches = (n + batch_size - 1) // batch_size
    imported = 0
    errors: list[str] = []

    for bi in range(num_batches):
        lo = bi * batch_size
        hi = min(lo + batch_size, n)
        chunk = data_rows[lo:hi]
        try:
            mapped = _call_llm_map_rows(headers, chunk, bi, num_batches)
        except Exception as exc:
            errors.append(f"第 {bi + 1} 批处理失败：{exc}")
            continue
        for rec in mapped:
            try:
                payload = _record_to_save_payload(rec)
                db_mod.save_student_record(payload)
                imported += 1
            except ValueError as ve:
                errors.append(str(ve))
            except Exception:
                log.exception("save student import row")
                errors.append("某行保存失败，请检查姓名/学号是否至少填一项。")

    return {
        "imported": imported,
        "totalRows": n,
        "batches": num_batches,
        "batchSize": batch_size,
        "errors": errors[:50],
    }
